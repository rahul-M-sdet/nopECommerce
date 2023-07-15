import openpyxl

def getRowCount(file,sheetName):
    workbook=openpyxl.load_workbook(".\\TestData\\PythonDemo.xlsx")
    sheet= workbook[sheetName]
    return (sheet.max_row)

def getColumnCount(file,sheetName):
    workbook=openpyxl.load_workbook(".\\TestData\\PythonDemo.xlsx")
    sheet= workbook[sheetName]
    return (sheet.max_column)

def readData(file,sheetName,rowNo, columnNo):
    workbook=openpyxl.load_workbook(".\\TestData\\PythonDemo.xlsx")
    sheet = workbook[sheetName]
    return sheet.cell(row=rowNo , column=columnNo).value

def writeData(file,sheetName,rowNo, columnNo):
    workbook=openpyxl.load_workbook(".\\TestData\\PythonDemo.xlsx")
    sheet = workbook[sheetName]
    return sheet.cell(row=rowNo , column=columnNo).value=data
    workbook.save()